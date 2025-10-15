#!/usr/bin/env python3
# BEGIN PATCH: add unescape import for HTML entity decoding
import os
import csv
import re
import json
import requests  # pip install requests
from openpyxl import Workbook
import pandas as pd  # indien later nodig
from tqdm import tqdm  # progress bar
import xml.etree.ElementTree as ET
import smtplib
from email.message import EmailMessage
from html import unescape
# Track duplicate log entries to avoid writing them multiple times
_LOGGED_DUPLICATES = set()

# Remove unsupported XML characters from text
def clean_xml_text(text):
    if text is None:
        return ""
    text = str(text) 
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", text)
    # Manually escape problematic XML characters
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return text

# Toggle UIC/STC HTTP POSTing: set to "YES" to enable POST calls, "NO" to disable
POST_UIC_STC = "NO"

# Helper for leading zero stripping
def strip_leading_zeros(s):
    return s.lstrip("0") if isinstance(s, str) else s

#########################
# PADINSTELLINGEN
#########################
SPEC_JSON_PATH = "SpecificationAttributeRespons.json"

# Base directory for input and output files - configureerbaar maken
BASE_DIR = os.path.dirname(os.path.realpath(__file__))
# Path naar inputbestanden (lokale override)
LOCAL_INPUT_DIR = "/Users/martijn/Desktop/IMG_PW/UPDATE/scripts"
# Detect available suffixes dynamically
def detect_suffixes():
    import glob
    files = glob.glob(os.path.join(LOCAL_INPUT_DIR, "art*in"))
    suffixes = []
    for file in files:
        name = os.path.basename(file)
        if name.startswith("art") and name.endswith("in"):
            mid = name[3:-2]
            if mid.isdigit():
                suffixes.append(mid + "in")
    return suffixes
# Helper to find file by prefix and suffix in BASE_DIR
def find_file_by_prefix(suffix, prefix):
    import glob
    pattern = os.path.join(LOCAL_INPUT_DIR, f"{prefix}*{suffix}")
    matches = glob.glob(pattern)
    return matches[0] if matches else None
# Alternatief: gebruik os.getcwd() voor huidige directory of via command line argument
# BASE_DIR = os.path.join(os.getcwd(), "VBPW0000")
# Global formula placeholder
GLOBAL_FORMULA = ""
GLOBAL_SUFFIX = ""

# per-run log file placeholders (set in run_conversion)
LOG_FILE = ""
ATK_LOG_FILE = ""

# ensure default log directory and files so LOG_FILE is never empty
DEFAULT_LOG_DIR = os.path.join(BASE_DIR, "converted-files")
os.makedirs(DEFAULT_LOG_DIR, exist_ok=True)
if not LOG_FILE:
    LOG_FILE = os.path.join(DEFAULT_LOG_DIR, "log.txt")
if not ATK_LOG_FILE:
    ATK_LOG_FILE = os.path.join(DEFAULT_LOG_DIR, "ATK-log.txt")

#########################
# GLOBALE DEFINITIES
#########################
EXPECTED_ATK = set()
for s in [
    "Afmetingen", "CT", "CT_ARMEN", "CT_COMFORT", "CT_DETAILS", "CT_FUNCTIONEEL", "CT_KLEUREN",
    "CT_POTEN", "CT_RELAX", "CT_ZITCOMFORT", "CT_ZITTING", "CT2", "LABEL", "Omschrijving",
    "pk_baenks", "Prijskaarttekst", "SHOWROOMMODEL", "Stofgroep 1", "Uitvoeringen"
]:
    EXPECTED_ATK.add(s.upper())
    EXPECTED_ATK.add(s.upper().replace(" ", ""))
PREFERRED_ATK_ORDER = ["OMSCHRIJVING", "CT", "UITVOERINGEN", "AFMETINGEN", "SHOWROOMMODEL"]

BASE_HEADERS = [
    "Artikelnummer", "Artikelomschrijving", "Barcode", "Eenh_aankoop", "Eenh_verkoop",
    "Bestelhoeveelheid", "nvt", "nvt", "nvt", "nvt", "nvt",
    "Standaardprijs verkoopprijs", "nvt", "Standaard inkoopprijs",
    "Standaard magazijn", "Standaard locatie", "Omzetgroepnummer", "Crediteurnummer",
    "Code in collectie", "Minimum voorraad", "Maximum voorraad", "Inkoopomschrijving",
    "Code materiaal", "Code uitvoering", "Groepscode", "Stofbreedte", "Patroonhoogte",
    "Commissiecode", "Aanvullende omschrijving", "Hoofdartikelnummer", "Hoofdartikelomschrijving",
    "Code sfeer", "Code concept", "Hoofdartikel afdelingnummer", "Code vrije omschrijving",
    "Code vrije prijs", "Code inkoopvaluta", "Adviesprijs", "nvt", "nvt", "nvt", "nvt",
    "Hoofdartikel korting 1", "Hoofdartikel korting 2", "Hoofdartikel opslag conversiefactor",
    "Hoofdartikel code collectie", "Catalogus uitvoering", "Catalogus soort", "kenmerken",
    "begindatum contract", "einddatum contract", "soort afronding", "prijs afronden op",
    "verzamelen order op \"0\"", "code catalogus op 1", "omschrijving catalogus",
    "Levertijd", "leverweek_CONV", "Artikel publicatie", "Artikelnr leverancier", "Volume", "Gewicht",
    "Cat_extra_detail", "Artikelsoort", "Aantal colli", "Hoogte", "Breedte", "Lengte",
    "Zithoogte", "Zitdiepte", "Hoogte onderzijde blad", "Hoogte armligger",
    "Hoogte verpakking", "Breedte verpakking", "Lengte verpakking",
    "Volume verpakking", "Aantal colli verpakking",
    "Artikel publicatie verkoop - NIEUW", "Voorraad houden - NIEUW", "ParentSKU"
]

def rename_nvt_headers(headers):
    nvt_counter = 1
    new_headers = []
    for header in headers:
        if header.lower() == "nvt":
            new_headers.append(f"nvt{nvt_counter}")
            nvt_counter += 1
        else:
            new_headers.append(header)
    return new_headers

BASE_HEADERS = rename_nvt_headers(BASE_HEADERS)

# Velden voor CONV-kolommen in het ART-bestand
convert_headers = [
    "Artikel publicatie",
    "Artikelsoort",
    "Catalogus soort",
    "Catalogus uitvoering",
    "Code concept",
    "Code in collectie",
    "Code materiaal",
    "Code sfeer",
    "Code uitvoering",  # speciale verwerking
    "Commissiecode",
    "Eenh_aankoop",
    "Eenh_verkoop",
    "Groepscode",
    "Hoofdartikel code collectie",
    "Omzetgroepnummer",
    "soort afronding",
]

def build_extra_conv_headers(convert_headers):
    result = []
    for hdr in convert_headers:
        up = hdr.upper()
        if up == "AFMETINGEN":
            result += ["AFMETINGEN_CONV", "PRIJSKAART AFMETINGEN_CONV"]
        elif up == "UITVOERINGEN":
            result += ["UITVOERINGEN_CONV", "UITVOERINGEN Webshop_CONV"]
        elif up == "CODE UITVOERING":
            result.append("Code uitvoering_CONV")
        else:
            result.append(f"{hdr}_CONV")
    return result

extra_conv_headers = build_extra_conv_headers(convert_headers)

def get_extra_atk_headers(extra_atk_fields):
    headers = []
    for field in extra_atk_fields:
        up = field.upper()
        if up == "AFMETINGEN":
            headers += ["ATK_Afmetingen_CONV", "ATK_Prijskaart Afmetingen_CONV"]
        elif up == "UITVOERINGEN":
            headers += ["ATK_Uitvoeringen_CONV", "ATK_Uitvoeringen Webshop_CONV"]
        else:
            headers.append(f"ATK_{field.capitalize()}_CONV")
    return headers

# Unieke suffix voor Catalogus-velden
CONV_HEADERS = []

#########################
# HULPFUNCTIES
#########################
def format_number(x):
    try:
        f = float(x)
        return str(int(f)) if f.is_integer() else str(f)
    except:
        return x


def clean_value(value):
    if isinstance(value, str):
        # DEBUG: Check if clean_value removes <br> tags for article 10108948
        if "10108948" in str(value) or ("<br>" in value and "gewoon" in value):
            print(f"\n🔍 DEBUG CLEAN_VALUE:")
            print(f"Input: '{value[:100]}...'")
            cleaned = "".join(c for c in value if ord(c) >= 32 or c in "\n\r")
            print(f"Output: '{cleaned[:100]}...'")
            if "<br>" in value and "<br>" not in cleaned:
                print("❌ clean_value REMOVED <br> tags!")
            elif "<br>" in cleaned:
                print("✅ clean_value preserved <br> tags")
            return cleaned
        else:
            return "".join(c for c in value if ord(c) >= 32 or c in "\n\r")
    return value

# Helper to clean ATK token values (strip quotes/whitespace at ends, keep internal)
def clean_atk_token(s: str) -> str:
    """
    Strip surrounding quotes and whitespace from an ATK value.
    Keeps internal quotes (e.g., inside HTML/script), only trims at the ends.
    """
    if not isinstance(s, str):
        return s
    # Trim whitespace at both ends
    s = s.strip()
    # Remove any number of leading/trailing straight quotes
    s = re.sub(r'^(?:"|\')+|(?:"|\')+$', '', s).strip()
    return s

def _count_trailing_ws(s: str) -> int:
    """
    Count trailing whitespace including normal spaces, tabs and non‑breaking spaces (\u00A0).
    """
    if not isinstance(s, str):
        return 0
    m = re.search(r"[ \t\u00A0]+$", s)
    return len(m.group(0)) if m else 0

def _merge_atk_segments(segs):
    """
    Combineer ATK-segmenten tot een string waarbij:
    - regels met 2 of meer spaties aan het einde een <br> krijgen
    - regels zonder spaties aan het einde en een korte volgende regel worden samengevoegd
    - alle inhoud intact blijft, inclusief exacte spacing
    """
    merged_parts = []
    i = 0
    while i < len(segs):
        current = segs[i].rstrip("\r\n")
        next_seg = segs[i+1] if i+1 < len(segs) else ""
        current_trailing_spaces = len(current) - len(current.rstrip(" \t\u00A0"))

        # Voeg exacte tekst toe zonder strippen
        line = current

        # Check op harde <br> door trailing spaties
        if current_trailing_spaces >= 2:
            line = current.rstrip(" \t\u00A0") + "<br>"
            merged_parts.append(line)
            i += 1
            continue

        # Voeg volgende regel toe als de huidige regel niet eindigt met spatie
        if i + 1 < len(segs):
            next_line = next_seg.rstrip("\r\n")
            next_clean = next_line.strip()
            next_trailing_spaces = len(next_line) - len(next_line.rstrip(" \t\u00A0"))
            if not current.endswith(" ") and len(next_clean) <= 2:
                # Samenvoegen van woorden zoals gewoo+n
                line += next_clean
                if next_trailing_spaces >= 2:
                    line += "<br>"
                merged_parts.append(line)
                i += 2
                continue

        # Default append
        merged_parts.append(line)
        i += 1

    # Alleen <br> toevoegen als er meer dan één regel is
    if len(segs) > 1:
        merged = "".join(merged_parts)
    else:
        merged = segs[0]
    # Extra normalisatie: vervang spatie-achtige tekens tussen woorden door gewone spatie
    merged = re.sub(r"(?<=\w)\s+(?=\w)", " ", merged)
    # Fix common character encoding issues
    merged = merged.replace("ï¿½", "")
    merged = merged.replace("Ã¯Â¿Â½", "")
    merged = merged.replace("Ã¡", "á")
    merged = merged.replace("Ã©", "é")
    merged = merged.replace("Ã­", "í")
    merged = merged.replace("Ã³", "ó")
    merged = merged.replace("Ãº", "ú")
    merged = merged.replace("Ã±", "ñ")
    merged = merged.replace("Ã¼", "ü")
    merged = merged.replace("Ã¶", "ö")
    merged = merged.replace("Ã¤", "ä")
    merged = merged.replace("< br>", "<br>").replace("</u l>", "</ul>")
    # Normaliseer expliciet niet-breekbare spaties naar gewone spatie
    merged = merged.replace("\u00A0", " ")
    # Fix encoding artifacts to proper characters
    merged = merged.replace("�", "")
    merged = merged.replace("Ã¡", "á").replace("Ã©", "é").replace("Ã­", "í").replace("Ã³", "ó").replace("Ãº", "ú")
    merged = merged.replace("Ã±", "ñ").replace("Ã¼", "ü").replace("Ã¶", "ö").replace("Ã¤", "ä")
    # Vervang "Laat�je�uitgebreid" soort gevallen met normale spaties
    merged = re.sub(r"[�\uFFFD\u00A0]+", " ", merged)
    # Fix: vervang rare spatie-achtige karakters tussen woorden
    # (regel verwijderd om HTML en spacing te behouden)
    return merged

#########################
# PROCESS-FUNCTIES
#########################
def process_art_data():
    rows = []
    indices = [11,13,37,44,56,59,60,63,64,65,66,67,68,69,70,71,72,73,74,75]
    if not os.path.exists(INPUT_PATH):
        print(f"Inputbestand niet gevonden: {INPUT_PATH}")
        return []
    with open(INPUT_PATH, "r", encoding="latin1", errors="ignore") as f:
        reader = csv.reader(f, delimiter=",", quotechar='"')
        for r in reader:
            r = [c.strip() for c in r]
            for i in indices:
                if i < len(r):
                    num = r[i].lstrip("0")
                    if num.startswith("."):
                        num = "0"+num
                    r[i] = num
            r.insert(57, "")  # placeholder leverweek_CONV
            rows.append(r)
    for r in rows:
        while len(r) < len(BASE_HEADERS):
            r.append("")
    # bereken leverweek_CONV per hoofdartikelnummer
    groep = {}
    for r in rows:
        key = r[29] if len(r)>29 else ""
        groep.setdefault(key, []).append(r)
    for grp in groep.values():
        unieke = {row[56] for row in grp if len(row)>56 and row[56]}
        try: sorted_lt = sorted(unieke, key=float)
        except: sorted_lt = sorted(unieke)
        if len(sorted_lt)>=2:
            conv = f"{format_number(sorted_lt[0])} / {format_number(sorted_lt[-1])}"
        elif len(sorted_lt)==1:
            conv = format_number(sorted_lt[0])
        else:
            conv = ""
        for row in grp:
            if len(row)>57: row[57]=conv
            row[-1] = ""  # ParentSKU vullen we later
    return rows

def process_atk_data():
    """
    Parse the ATK input file line-by-line to preserve embedded quotes and HTML,
    then merge segments into a single string per field using <br> only when
    multiple segments exist.
    """
    data = {}
    fields = set()
    if not os.path.exists(ATK_INPUT_PATH):
        print(f"ATK bestand niet gevonden: {ATK_INPUT_PATH}")
        return data, fields
    with open(ATK_INPUT_PATH, "r", encoding="latin1", errors="replace") as f:
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            parts = re.split(r'","', line)
            parts = [p.strip('"') for p in parts]
            if len(parts) < 8:
                continue
            # Extract article number and field name
            art = parts[0].lstrip('"').strip()
            fld = parts[2].strip().upper()
            fields.add(fld)
            # Rejoin everything after the 7th split as the raw value
            val = '"'.join(parts[7:])
            data.setdefault(art, {}).setdefault(fld, []).append(val)
    # Merge segments per art/field
    for art, fl in data.items():
        for fld, segs in fl.items():
            merged = _merge_atk_segments(segs)
            # Strip trailing spaces from single-line fields
            if len(segs) == 1:
                merged = merged.rstrip()
            # DEBUG: Check if <br> tags are being created for article 10108948
            if art.strip() == "10108948" and fld == "CT":
                print(f"\n🔍 DEBUG AFTER ATK MERGE for article {art}, field {fld}:")
                print(f"Input segments: {segs[:3]}")
                print(f"Merged result: '{merged[:150]}...'")
                if "<br>" in merged:
                    print("✅ <br> tags are created correctly in ATK processing")
                else:
                    print("❌ <br> tags are missing after ATK merge")
            # Herstel spatie na sluit-haakje voor cijfers en schrijf weg
            merged = re.sub(r'\)(\d)', r') \1', merged)
            data[art][fld] = merged
    return data, sorted(fields)

def determine_atk_fields(unique_fields, atk_data):
    KEEP_ATK_FIELDS = {
        'AFMETINGEN', 'CT', 'CT_ARMEN', 'CT_COMFORT', 'CT_DETAILS', 'CT_FUNCTIONEEL',
        'CT_KLEUREN', 'CT_POTEN', 'CT_RELAX', 'CT_ZITCOMFORT', 'CT_ZITTING', 'CT2',
        'LABEL', 'OMSCHRIJVING', 'PK_BAENKS', 'PRIJSKAARTTEKST', 'SHOWROOMMODEL',
        'STOFGROEP1', 'UITVOERINGEN'
    }
    # Build lists of present fields in preferred order, then other present, then always-include missing
    present_pref = [f for f in PREFERRED_ATK_ORDER if f in unique_fields and f in KEEP_ATK_FIELDS]
    present_rest = sorted([f for f in unique_fields if f in KEEP_ATK_FIELDS and f not in PREFERRED_ATK_ORDER])
    missing = sorted(f for f in KEEP_ATK_FIELDS if f not in unique_fields)
    result_fields = present_pref + present_rest + missing

    # Log any unexpected fields
    unexpected = [f for f in unique_fields if f not in KEEP_ATK_FIELDS]
    field_articles = {f: [] for f in unexpected}
    for artnr, fields_map in atk_data.items():
        for f in unexpected:
            if f in fields_map:
                field_articles[f].append(artnr)
    with open(ATK_LOG_FILE, "a", encoding="utf-8") as log:
        if unexpected:
            log.write("Unexpected ATK fields:\n")
            for f, arts in field_articles.items():
                log.write(f"{f}: {', '.join(arts)}\n")
        else:
            log.write("All ATK fields are expected\n")

    return result_fields

def process_mvm_file():
    blocks = {}
    try:
        with open(MVM_INPUT_PATH, "r", encoding="windows-1252", errors="strict", newline="") as f:
            reader = csv.reader(f, delimiter=",", quotechar='"')
            for r in reader:
                if len(r) < 3:
                    continue
                key, txt = r[0].strip(), r[2]
                blocks.setdefault(key, []).append(txt)
    except Exception as e:
        print(f"Fout bij lezen MVM: {e}")
        return {}
    result = {}
    MAX = 32767
    for k, lst in blocks.items():
        # Just keep the list of lines (do NOT join to string)
        result[k] = lst
    return result

def process_uic_data():
    d = {}
    if not os.path.exists(UIC_INPUT_PATH):
        print(f"UIC niet gevonden: {UIC_INPUT_PATH}")
        return d
    with open(UIC_INPUT_PATH,"r",encoding="utf-8",errors="replace",newline="") as f:
        for r in csv.reader(f,delimiter=",",quotechar='"'):
            if len(r)<2: continue
            d[r[0].strip()] = r[1].strip()
    return d

def process_stc_data():
    d = {}
    if not os.path.exists(STC_INPUT_PATH):
        print(f"STC niet gevonden: {STC_INPUT_PATH}")
        return d
    with open(STC_INPUT_PATH,"r",encoding="utf-8",errors="replace",newline="") as f:
        for r in csv.reader(f,delimiter=",",quotechar='"'):
            if len(r)<2: continue
            d[r[0].strip()] = r[1].strip()
    return d


def process_uic_and_stc():
    """
    Read uic0000 and stc0000 files for the current suffix and post any missing
    code/name pairs to KatanaPIM. Log creations and mismatches to uic-stc-log.txt.
    """
    spec_map = process_specifications()
    tasks = [
        ("Catalogus uitvoering", UIC_INPUT_PATH, "157"),
        ("Catalogus soort",      STC_INPUT_PATH, "158"),
    ]
    # clear or create uic-stc-log.txt
    open("uic-stc-log.txt", "w", encoding="utf-8").close()
    for attr_name, input_path, attr_id in tasks:
        if not input_path:
            with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                log.write(f"Bestand niet gevonden voor {attr_name}, overslaan.\n")
            continue
        with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
            log.write(f"--- Processing {attr_name} from {input_path}\n")
        posted_count = 0
        skipped_count = 0
        rows = []
        if os.path.exists(input_path):
            with open(input_path, "r", encoding="latin1", errors="ignore") as f:
                reader = csv.reader(f, delimiter=",", quotechar='"')
                for r in reader:
                    if len(r) >= 2:
                        code = r[0].strip()
                        orig_name = r[1].strip()
                        if not code:
                            code = orig_name
                        # build full option name, trimming any trailing spaces
                        name = f"{code} - {orig_name}".rstrip()
                        rows.append((code, name))
        else:
            with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                log.write(f"File not found: {input_path}\n")
            continue
        for code, name in tqdm(rows, desc=f"Posting {attr_name}"):
            options = spec_map.get(attr_name, {})
            # build case-insensitive key map for existing options
            lower_options_map = {k.lower(): k for k in options.keys()}
            stripped = code.lstrip("0")
            if POST_UIC_STC.upper() == "NO":
                # determine if option already exists (case-insensitive)
                existing_key = None
                if code in options:
                    existing_key = code
                elif code.lower() in lower_options_map:
                    existing_key = lower_options_map[code.lower()]
                elif stripped in options:
                    existing_key = stripped
                elif stripped.lower() in lower_options_map:
                    existing_key = lower_options_map[stripped.lower()]
                if existing_key is None:
                    with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                        log.write(f"Missing {attr_name} [{code}] -> \"{name}\" (posting disabled)\n")
                else:
                    existing = options.get(existing_key)
                    if existing != name:
                        with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                            log.write(f"Name mismatch for {attr_name} [{code}]: PIM has \"{existing}\", file has \"{name}\"\n")
                    skipped_count += 1
                continue
            # determine if option already exists (case-insensitive)
            existing_key = None
            if code in options:
                existing_key = code
            elif code.lower() in lower_options_map:
                existing_key = lower_options_map[code.lower()]
            elif stripped in options:
                existing_key = stripped
            elif stripped.lower() in lower_options_map:
                existing_key = lower_options_map[stripped.lower()]

            # only create new options if not existing
            if existing_key is None:
                with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                    log.write(f"Creating new {attr_name} [{code}] -> \"{name}\"\n")
                payload = {"name": name, "code": code, "subTitle": "", "displayOrder": 0}
                url = f"https://img.katanapim.com/api/v2/specificationAttributes/{attr_id}/specificationAttributeOptions"
                resp = requests.post(url, headers={
                    "apikey": "69d88bd8-08b6-4885-bcdc-e900c7b564c5",
                    "Content-Type": "application/json"
                }, json=payload)
                with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                    if resp.status_code in (200, 201):
                        log.write(f"POSTED \"{name}\" in KatanaPIM\n")
                        options[code] = name
                        options[stripped] = name
                        posted_count += 1
                    elif resp.status_code == 400 and "already exists" in resp.text:
                        # option already exists; nothing to do
                        options[code] = name
                        options[stripped] = name
                        skipped_count += 1
                    else:
                        log.write(f"FAILED create {attr_name} [{code}]: {resp.status_code} {resp.text}\n")
            else:
                # already exists (exact or case-insensitive), skip creation
                with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
                    log.write(f"Skipping existing {attr_name} [{code}] -> \"{name}\"\n")
                skipped_count += 1

        with open("uic-stc-log.txt", "a", encoding="utf-8") as log:
            log.write(f"{attr_name}: {posted_count} posted, {skipped_count} existing\n")

    # after posting, force a refresh of the specification JSON so new options are picked up
    if POST_UIC_STC.upper() == "YES":
        try:
            os.remove(SPEC_JSON_PATH)
        except OSError:
            pass

def process_specifications():
    global LOG_FILE
    if not LOG_FILE:
        LOG_FILE = os.path.join(DEFAULT_LOG_DIR, "log.txt")
    m = {}
    if not os.path.exists(SPEC_JSON_PATH):
        url = "https://img.katanapim.com/api/v1/Specifications?PageSize=150"
        headers_req = {"apikey":"69d88bd8-08b6-4885-bcdc-e900c7b564c5"}
        try:
            resp = requests.get(url, headers=headers_req)
            resp.raise_for_status()
            data = resp.json()
            with open(SPEC_JSON_PATH,"w",encoding="utf-8") as f:
                json.dump(data,f,ensure_ascii=False,indent=4)
        except Exception as e:
            print(f"Fout bij API: {e}")
            return m
    else:
        try:
            with open(SPEC_JSON_PATH,"r",encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"Fout bij openen JSON: {e}")
            return m
    conv_set = {h.lower() for h in convert_headers}
    conv_set.add('gp_hoofdartikel_afdelingsnummer')
    conv_set.add('gp_soort_afronding')
    for item in data.get("Items", []):
        code = item.get("Code", "").strip().lower()
        if code in conv_set:
            # Phase 1: collect raw_map with duplicate detection among raw codes only
            raw_map = {}
            for opt in item.get("Options", []):
                c = opt.get("Code", "").strip()
                n = opt.get("Name", "").strip()
                if c in raw_map and raw_map[c] != n:
                    msg = f"Dubbele {item.get('Code')} code {c}: '{raw_map[c]}' vs '{n}'"
                    if msg not in _LOGGED_DUPLICATES:
                        with open(LOG_FILE, "a", encoding="utf-8") as log:
                            log.write(msg + "\n")
                        _LOGGED_DUPLICATES.add(msg)
                raw_map[c] = n
            # Phase 2: extend opt_map with zero-stripped and padded keys
            opt_map = {}
            for c, n in raw_map.items():
                opt_map[c] = n
                opt_map[strip_leading_zeros(c)] = n
                # Also map padded two-digit codes for single-digit numeric codes
                if re.fullmatch(r'\d', c):
                    opt_map[c.zfill(2)] = n
            m[item.get("Code", "").strip()] = opt_map
    # Alias for soort afronding header lookup
    if 'gp_soort_afronding' in m:
        m['soort afronding'] = m['gp_soort_afronding']
    return m

def create_art_workbook(rows_list, atk_data, extra_atk_fields):
    spec_map = process_specifications()
    posted_codes = set()
    extra_atk_conv_headers = get_extra_atk_headers(extra_atk_fields)
    # Always include Baenks-specific ATK CONV fields for all formulas
    bs_always_atk_conv = [
        "ATK_Ct_armen_CONV",
        "ATK_Ct_comfort_CONV",
        "ATK_Ct_details_CONV",
        "ATK_Ct_functioneel_CONV",
        "ATK_Ct_kleuren_CONV",
        "ATK_Ct_poten_CONV",
        "ATK_Ct_relax_CONV",
        "ATK_Ct_zitcomfort_CONV",
        "ATK_Ct_zitting_CONV",
        "ATK_Pk_baenks_CONV",
        "ATK_Prijskaarttekst_CONV",
        "ATK_Stofgroep1_CONV",
    ]
    for col in bs_always_atk_conv:
        if col not in extra_atk_conv_headers:
            extra_atk_conv_headers.append(col)

    # Build mapping from convert_headers to their corresponding CONV columns
    conv_map = {}
    for hdr, conv_hdr in zip(convert_headers, extra_conv_headers):
        up = hdr.upper()
        if up == "AFMETINGEN":
            conv_map[hdr] = ["AFMETINGEN_CONV", "PRIJSKAART AFMETINGEN_CONV"]
        elif up == "UITVOERINGEN":
            conv_map[hdr] = ["UITVOERINGEN_CONV", "UITVOERINGEN Webshop_CONV"]
        elif up == "CODE UITVOERING":
            conv_map[hdr] = ["Code uitvoering_CONV"]
        else:
            conv_map[hdr] = [f"{hdr}_CONV"]

    # Rebuild HEADERS_EXT with interleaving
    HEADERS_EXT = []
    for h in BASE_HEADERS:
        HEADERS_EXT.append(h)
        # case-insensitive match for conv_map keys
        key = h if h in conv_map else h.lower() if h.lower() in conv_map else None
        if key:
            HEADERS_EXT.extend(conv_map[key])
    HEADERS_EXT += extra_atk_conv_headers
    # NEW: concatenated price card description column
    HEADERS_EXT.append("prijskaart_omschrijving")
    HEADERS_EXT += [h+"_2" for h in CONV_HEADERS]
    # Add SpecificatieGroepen header
    HEADERS_EXT.append("SpecificatieGroepen")
    # Add GrandParentExternalKey header
    HEADERS_EXT.append("GrandParentExternalKey")

    # log ATK-velden
    with open(LOG_FILE,"a",encoding="utf-8") as log:
        actual = {f.upper() for f in extra_atk_fields}
        diff = actual.symmetric_difference(EXPECTED_ATK)
        if diff:
            log.write("Waarschuwing ATK-velden mismatch: "+", ".join(diff)+"\n")
        else:
            log.write("Alle ATK-velden OK\n")

    wb = Workbook(); ws = wb.active
    # Prepare to collect all cleaned output rows for splitting
    all_out_rows = []
    ws.append(HEADERS_EXT)
    # Build a map of header to column index for transformations
    header_idx_map = {h: i for i, h in enumerate(HEADERS_EXT)}

    uic = process_uic_data()
    stc = process_stc_data()
    idx_map = {h:i for i,h in enumerate(BASE_HEADERS)}
    # Find index of ATK_Showroommodel_CONV header for later use
    try:
        showroom_idx = HEADERS_EXT.index("ATK_Showroommodel_CONV")
    except ValueError:
        showroom_idx = None

    for art_row in tqdm(rows_list, desc="Processing ART rows"):
        artnr = art_row[0].strip()
        if not artnr or not art_row[1].strip():
            with open(LOG_FILE,"a",encoding="utf-8") as log:
                log.write(f"[SKIP] verplicht missing voor {artnr}\n")
            continue

        out = []
        # For ParentSKU we need to know the Code uitvoering_CONV value, so track it
        code_uitvoering_conv_val = None
        for h in BASE_HEADERS:
            base_val = art_row[idx_map[h]] if h in idx_map and idx_map[h] < len(art_row) else ""
            out.append(base_val)
            if h in conv_map:
                cmap = spec_map.get(h, {})
                # Special handling for "Code uitvoering" with padded lookup and logging
                if h == "Code uitvoering":
                    # Always pad single-digit codes to two digits (including "0" -> "00")
                    padded_input = base_val.zfill(2) if re.fullmatch(r'\d', base_val or "") else base_val
                    # override the base value cell with the padded code
                    out[-1] = padded_input
                    # perform lookup using the padded code first, then fallback to original
                    cv = None
                    if padded_input in cmap:
                        cv = cmap[padded_input]
                    elif base_val in cmap:
                        cv = cmap[base_val]
                    else:
                        # Log missing mapping to uic-stc-log.txt
                        with open("uic-stc-log.txt", "a", encoding="utf-8") as logf:
                            logf.write(f"Missing mapping for Code uitvoering: '{base_val}' (padded: '{padded_input}')\n")
                        cv = padded_input
                    out.append(cv)
                    code_uitvoering_conv_val = cv
                    continue  # skip generic conv_map branch for this header
                else:
                    for conv_col in conv_map[h]:
                        strip_val = strip_leading_zeros(base_val)
                        # prepare case-insensitive lookup map
                        lower_map = {k.lower(): v for k, v in cmap.items()}
                        cv = cmap.get(base_val)
                        if cv is None:
                            cv = cmap.get(strip_val)
                        if cv is None:
                            cv = lower_map.get(base_val.lower())
                        if cv is None:
                            cv = lower_map.get(strip_val.lower())
                        if cv is None:
                            cv = base_val
                        out.append(cv)

        # Set ParentSKU in the output row based on header_idx_map and code_uitvoering_conv_val
        ps_idx = header_idx_map.get("ParentSKU")
        if ps_idx is not None and code_uitvoering_conv_val:
            hoofdartikelnummer_val = art_row[idx_map["Hoofdartikelnummer"]].strip() if "Hoofdartikelnummer" in idx_map else artnr
            out[ps_idx] = f"{hoofdartikelnummer_val} - {code_uitvoering_conv_val}"

        # 3) ATK-velden - data is al correct verwerkt in process_atk_data()
        vals_atk = []
        for fld in extra_atk_fields:
            up = fld.upper()
            val = atk_data.get(artnr, {}).get(fld, "")
            # ATK fields should not be passed through clean_value, <br> must be preserved
            if isinstance(val, str):
                val = val.replace("\r", "").replace("\n", " ")
            # DEBUG: Check if <br> tags survive to Excel export for article 10108948
            if artnr.strip() == "10108948" and fld == "CT":
                print(f"\n🔍 DEBUG IN EXCEL EXPORT for article {artnr}, field {fld}:")
                print(f"ATK value from data: '{val[:150]}...'")
                if "<br>" in val:
                    print("✅ <br> tags survive to Excel export")
                else:
                    print("❌ <br> tags missing in Excel export - they were removed somewhere!")
            if up == "AFMETINGEN":
                full_val = val.strip()
                if "|" in full_val:
                    before_pipe, _ = full_val.split("|", 1)
                    # SWAP: ATK_Afmetingen_CONV gets before_pipe, ATK_Prijskaart Afmetingen_CONV gets full_val
                    vals_atk += [before_pipe.strip(), full_val]
                else:
                    vals_atk += [full_val, full_val]
            elif up == "UITVOERINGEN":
                if "|" in val:
                    a, b = val.split("|", 1)
                    vals_atk += [a, b]
                else:
                    vals_atk += [val, ""]
            else:
                vals_atk.append(val)

        # Avoid cleaning ATK fields below
        cleaned_atk_vals = vals_atk  # do not apply clean_value to ATK
        out.extend(vals_atk)

        # NEW: prijskaart_omschrijving = ATK_Uitvoeringen_CONV + ' ' + ATK_Afmetingen_CONV
        u_idx = header_idx_map.get("ATK_Uitvoeringen_CONV")
        p_idx = header_idx_map.get("ATK_Afmetingen_CONV")
        u_txt = out[u_idx].strip() if u_idx is not None and u_idx < len(out) and isinstance(out[u_idx], str) else ""
        p_txt = out[p_idx].strip() if p_idx is not None and p_idx < len(out) and isinstance(out[p_idx], str) else ""
        if u_txt and p_txt:
            combo = f"{u_txt} {p_txt}"
        else:
            combo = u_txt or p_txt
        out.append(combo)

        # Build SpecificatieGroepen with proper labels (always on child rows)
        categories = [
            "algemeen_child",
            "logistiek_child",
            "prijzen_child",
            "marketing_child",
            "teksten_child",
            "prijskaarten_child",
        ]
        # Suffix-specific child groups
        suf = GLOBAL_SUFFIX.lower()
        if suf == 'pw':
            categories.append("labels_pronto_wonen_child")
        elif suf == 'pm':
            categories.append("labels_profijt_meubel_child")
        if suf == 'bs':
            categories.append("teksten_child_baenks")
        spec_groups = "|".join(categories)
        out.append(spec_groups)

        # --- DEBUG OUTPUT FOR ARTIKEL 10108948 ---
        # Schrijf debug-output weg voor artikel 10108948
        if artnr.strip() == "10108948":
            try:
                ct_idx = header_idx_map.get("ATK_Ct_CONV")
                debug_val = out[ct_idx] if ct_idx is not None and ct_idx < len(out) else ""
                with open("debug-output-10108948.txt", "w", encoding="utf-8") as f:
                    f.write(debug_val)
            except Exception as e:
                print(f"Fout bij schrijven debug-output voor 10108948: {e}")
        # --- END DEBUG OUTPUT ---

        # GrandParentExternalKey: combine Hoofdartikelnummer and Hoofdartikelomschrijving
        hp_idx = idx_map["Hoofdartikelnummer"]
        ho_idx = idx_map["Hoofdartikelomschrijving"]
        hoofnr = art_row[hp_idx].strip() if hp_idx < len(art_row) else ""
        hoofom = art_row[ho_idx].strip() if ho_idx < len(art_row) else ""
        out.append(f"{hoofnr} - {hoofom}")

        # For all decimal fields except standard sales price, replace dot with comma
        comma_decimal_cols = ["Adviesprijs", "Hoofdartikel opslag conversiefactor", "Volume"]
        for col in comma_decimal_cols:
            idx = header_idx_map.get(col)
            if idx is not None and idx < len(out) and out[idx] != "":
                out[idx] = out[idx].replace(".", ",")
        # Ensure standard sales price uses dot as decimal separator
        sp_idx = header_idx_map.get("Standaardprijs verkoopprijs")
        if sp_idx is not None and sp_idx < len(out) and out[sp_idx] != "":
            out[sp_idx] = out[sp_idx].replace(",", ".")

        # Strip leading zeros on integer fields
        int_strip_cols = ["Bestelhoeveelheid", "Standaard magazijn", "Minimum voorraad", "Maximum voorraad", "prijs afronden op", "Stofbreedte", "Patroonhoogte"]
        for col in int_strip_cols:
            idx = header_idx_map.get(col)
            if idx is not None and idx < len(out) and out[idx] != "":
                stripped = out[idx].lstrip("0")
                out[idx] = stripped if stripped != "" else "0"

        # Convert dimension decimals to comma
        dim_cols = ["Gewicht", "Hoogte", "Breedte", "Lengte", "Zithoogte", "Zitdiepte", "Hoogte onderzijde blad", "Hoogte verpakking", "Breedte verpakking", "Lengte verpakking", "Volume verpakking", "Hoogte armligger"]
        for col in dim_cols:
            idx = header_idx_map.get(col)
            if idx is not None and idx < len(out) and out[idx] != "":
                out[idx] = out[idx].replace(".", ",")

        cleaned = []
        for i, c in enumerate(out):
            colname = HEADERS_EXT[i]
            if colname.startswith("ATK_") and isinstance(c, str):
                cleaned.append(c)  # skip clean_value
            else:
                cleaned.append(clean_value(c))
        # DEBUG: Check final cleaned row for article 10108948
        if artnr.strip() == "10108948":
            ct_idx = None
            try:
                ct_idx = header_idx_map.get("ATK_Ct_CONV")
                if ct_idx and ct_idx < len(cleaned):
                    print(f"\n🔍 DEBUG FINAL CLEANED ROW for article {artnr}:")
                    print(f"ATK_Ct_CONV value: '{str(cleaned[ct_idx])[:150]}...'")
                    if "<br>" in str(cleaned[ct_idx]):
                        print("✅ <br> tags are in final Excel row")
                    else:
                        print("❌ <br> tags missing from final Excel row")
            except:
                pass
        all_out_rows.append(cleaned)

    # --- SWAP ATK_Afmetingen_CONV and ATK_Prijskaart Afmetingen_CONV just before writing ---
    try:
        idx_a = HEADERS_EXT.index("ATK_Afmetingen_CONV")
        idx_b = HEADERS_EXT.index("ATK_Prijskaart Afmetingen_CONV")
        for row in all_out_rows:
            row[idx_a], row[idx_b] = row[idx_b], row[idx_a]
    except ValueError:
        pass

    for cleaned in all_out_rows:
        ws.append(cleaned)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"ART weggeschreven naar: {OUTPUT_PATH}")


def create_gp_workbook(rows_list, mvm_data):
    idx_map = {h: i for i, h in enumerate(BASE_HEADERS)}
    raw_map = {r[29].strip(): r for r in rows_list}

    # --- Build extra_atk from mvm0000 file for memotekst ---
    # Structure: extra_atk = { "WIELTJES": [ "WIELTJES       ,01,PRODUCTINFORMATIE...", ... ], ... }
    extra_atk = {}
    try:
        mvm_path = MVM_INPUT_PATH
        if os.path.exists(mvm_path):
            with open(mvm_path, "r", encoding="windows-1252", errors="replace") as f:
                for line in f:
                    line = line.rstrip("\r\n")
                    parts = line.split(",", 2)
                    if len(parts) >= 3:
                        key = parts[0].strip()
                        extra_atk.setdefault(key, []).append(line)
    except Exception as e:
        print(f"Fout bij lezen mvm0000-bestand voor extra_atk: {e}")
    # Build mapping from Crediteurnummer to supplier name
    lev_rows = process_lev_file()
    lev_map = {}
    for name, code, _ in lev_rows:
        raw_code = code.strip()
        stripped_code = strip_leading_zeros(raw_code)
        # Map both exact code and stripped code to supplier name
        lev_map[raw_code] = name
        lev_map[stripped_code] = name
    # Capture indices for korting fields
    idx_k1 = BASE_HEADERS.index("Hoofdartikel korting 1")
    idx_k2 = BASE_HEADERS.index("Hoofdartikel korting 2")
    # maak extra_dict voor contract, sfeer e.d.
    extra = {}
    for r in rows_list:
        key = r[29].strip() if len(r)>29 else ""
        if key and key not in extra:
            bd = r[49] if len(r)>49 else ""
            ed = r[50] if len(r)>50 else ""
            cs = r[31] if len(r)>31 else ""
            hoa= r[44] if len(r)>44 else ""
            hcc= r[45] if len(r)>45 else ""
            cc = r[32] if len(r)>32 else ""
            sa = r[51] if len(r)>51 else ""
            k1_val = r[idx_k1] if len(r)>idx_k1 else ""
            k2_val = r[idx_k2] if len(r)>idx_k2 else ""
            extra[key]=[bd,ed,cs,hoa,hcc,cc,sa,k1_val,k2_val]

    gp = {}
    for r in rows_list:
        if len(r)>57:
            sku = f"{r[29]} - {r[30]}"
            lvl = r[57]
            try:
                vals = set(float(x.strip()) for x in lvl.split("/") if x.strip())
            except:
                vals=set()
            if sku in gp:
                gp[sku]["leverweeks"].update(vals)
            else:
                gp[sku]={"SKU":sku, "Hoofdartikelnummer":r[29], "Hoofdartikelomschrijving":r[30], "leverweeks":vals}

    gp_rows=[]
    for sku, data in gp.items():
        # retrieve extension data for this Hoofdartikelnummer before using it
        ext = extra.get(data["Hoofdartikelnummer"].strip(), [""] * 9)
        lw = data["leverweeks"]
        if lw:
            s = sorted(lw)
            lever = f"{format_number(s[0])} / {format_number(s[-1])}" if len(s) >= 2 else format_number(s[0])
        else:
            lever = ""
        # Format memo field according to new instructions
        hoofdartnr = data["Hoofdartikelnummer"].strip().upper()
        raw_memo_lines = mvm_data.get(hoofdartnr)
        if isinstance(raw_memo_lines, list):
            memo = "\n".join(
                line.replace("\u00A0", " ") for line in raw_memo_lines
            )
        else:
            memo = str(raw_memo_lines or "").replace("\u00A0", " ")
        memo = unescape(memo)  # decode HTML entities
        # herstel encodingfouten: nu via encode/decode latin1->windows-1252 fallback
        memo = memo.encode("latin1", errors="replace").decode("windows-1252", errors="replace")
        # Fix diakritische tekens in memo
        memo = memo.replace("Ã¡", "á").replace("Ã©", "é").replace("Ã­", "í").replace("Ã³", "ó").replace("Ãº", "ú")
        memo = memo.replace("Ã±", "ñ").replace("Ã¼", "ü").replace("Ã¶", "ö").replace("Ã¤", "ä")
        memo = memo.replace("â€™", "’").replace("â€“", "–").replace("â€œ", "“").replace("â€", "”")
        memo = memo.replace("â€˜", "‘").replace("â€¢", "•").replace("â€¦", "…")
        formatted_memo = memo
        file_lower = os.path.basename(INPUT_PATH).lower()
        # Determine GP formula per Code concept for INHOUSE/House of Dutchz
        if GLOBAL_SUFFIX.lower() == 'in':
            code_concept = ext[5]
            if code_concept == '110':
                formule = 'House of Dutchz'
            else:
                formule = GLOBAL_FORMULA
        else:
            formule = GLOBAL_FORMULA
        spec = process_specifications()
        hcc_conv = spec.get("Hoofdartikel code collectie", {}).get(ext[4], ext[4])
        cc_conv = spec.get("Code concept", {}).get(ext[5], ext[5])
        raw = raw_map.get(data["Hoofdartikelnummer"], [])
        afdn = raw[idx_map["Hoofdartikel afdelingnummer"]].strip() if raw and idx_map["Hoofdartikel afdelingnummer"] < len(raw) else ""
        conv_afdn = spec.get("gp_hoofdartikel_afdelingsnummer", {}).get(strip_leading_zeros(afdn), afdn)
        gp_rows.append([
            data["SKU"], data["SKU"],
            data["Hoofdartikelnummer"], data["Hoofdartikelomschrijving"],
            formatted_memo, lever, formule, "algemeen_grandparent",
            ext[0], ext[1], ext[2],
            afdn, conv_afdn, ext[3],
            ext[7], ext[8],
            hcc_conv, cc_conv, ext[6]
        ])

    headers = [
        "SKU","ParentSKU","Hoofdartikelnummer","Hoofdartikelomschrijving",
        "Memoveld","Leverweek","formule","SpecificatieGroepen",
        "Begindatum contract","Einddatum contract","Code sfeer","Hoofdartikel afdelingnummer","Hoofdartikel afdelingnummer_CONV","Hoofdartikel opslag conversiefactor",
        "Hoofdartikel korting 1","Hoofdartikel korting 2",
        "Hoofdartikel code collectie_CONV","Code concept_CONV","Soort afronding",
        "Leveranciersnummer", "LeverancierNaam"
    ]
    wb=Workbook(); ws=wb.active
    ws.append(headers)
    # Find indices for fields that need decimal formatting
    # "Hoofdartikel opslag conversiefactor" is at index 12 in new headers
    opslag_idx = 12
    try:
        korting1_idx = headers.index("Hoofdartikel korting 1")
    except ValueError:
        korting1_idx = None
    try:
        korting2_idx = headers.index("Hoofdartikel korting 2")
    except ValueError:
        korting2_idx = None
    # Map crediteurnummer (with leading zeros stripped) to supplier name
    all_gp_rows = []
    for row in tqdm(gp_rows, desc="Building GP data"):
        # Copy Hoofdartikel korting values from the original ART rows
        raw = raw_map.get(row[2], [])
        if raw:
            if korting1_idx is not None and idx_k1 < len(raw):
                row[korting1_idx] = strip_leading_zeros(raw[idx_k1])
            if korting2_idx is not None and idx_k2 < len(raw):
                row[korting2_idx] = strip_leading_zeros(raw[idx_k2])
        # Append supplier code and supplier name based on Crediteurnummer from the original ART row
        cred = raw[idx_map["Crediteurnummer"]].strip() if raw and "Crediteurnummer" in idx_map else ""
        supplier_code = cred  # Dit is al eerder gedefinieerd
        supplier_name = lev_map.get(cred, lev_map.get(strip_leading_zeros(cred), ""))
        row.append(supplier_code)
        row.append(supplier_name)
        # Format "Hoofdartikel opslag conversiefactor" with comma
        if len(row) > opslag_idx and row[opslag_idx]:
            row[opslag_idx] = row[opslag_idx].replace('.', ',')
        # Also format "Hoofdartikel korting 1" and "Hoofdartikel korting 2" if present in headers
        if korting1_idx is not None and len(row) > korting1_idx and row[korting1_idx]:
            val = strip_leading_zeros(row[korting1_idx])
            row[korting1_idx] = val.replace('.', ',')
        if korting2_idx is not None and len(row) > korting2_idx and row[korting2_idx]:
            val = strip_leading_zeros(row[korting2_idx])
            row[korting2_idx] = val.replace('.', ',')
        cleaned_row = [clean_value(c) for c in row]
        ws.append(cleaned_row)
        all_gp_rows.append(cleaned_row)
    # Ensure output directory exists
    os.makedirs(os.path.dirname(GP_OUTPUT_PATH), exist_ok=True)
    wb.save(GP_OUTPUT_PATH)
    print(f"GP weggeschreven naar: {GP_OUTPUT_PATH}")




    # --- Export JSON voor memovelden ---
    import json
    json_out_path = GP_OUTPUT_PATH.replace("art0000", "art").replace("_GP.xlsx", "_GP-memovelden.json")
    memo_data = []
    for row in all_gp_rows:
        artikelnummer = row[0]
        hoofdnr = row[2]
        lines = mvm_data.get(hoofdnr.strip().upper(), [])
        if isinstance(lines, list) and lines:
            memo_txt = "<PRE>" + "\n".join(lines) + "\n</PRE>"
        else:
            memo_txt = ""
        memo_txt = memo_txt.encode("windows-1252", errors="replace").decode("windows-1252", errors="replace")
        memo_txt = memo_txt.replace("Ã¡", "á").replace("Ã©", "é").replace("Ã­", "í").replace("Ã³", "ó").replace("Ãº", "ú")
        memo_txt = memo_txt.replace("Ã±", "ñ").replace("Ã¼", "ü").replace("Ã¶", "ö").replace("Ã¤", "ä")
        memo_txt = memo_txt.replace("â€™", "’").replace("â€“", "–").replace("â€œ", "“").replace("â€", "”")
        memo_txt = memo_txt.replace("â€˜", "‘").replace("â€¢", "•").replace("â€¦", "…")
        # Vervang name door hoofdartikelnummer (row[2])
        name = row[2]  # Hoofdartikelnummer
        if memo_txt.strip():
            memo_data.append({"Artikelnummer": artikelnummer, "name": name, "Memoveld": memo_txt})
    with open(json_out_path, "w", encoding="utf-8") as jf:
        json.dump({"product": memo_data}, jf, ensure_ascii=False, indent=2)
    print(f"GP JSON memovelden geëxporteerd naar: {json_out_path}")
def process_lev_file():
    # use the suffix for dynamic lev file selection
    lev_path = os.path.join(BASE_DIR, f"lev0000{GLOBAL_SUFFIX}")
    if not os.path.exists(lev_path):
        print(f"LEV bestand niet gevonden: {lev_path}")
        return []
    rows = []
    with open(lev_path, "r", encoding="latin1", errors="ignore", newline="") as f:
        reader = csv.reader(f, delimiter=",", quotechar='"')
        for r in reader:
            if len(r) < 18:
                continue
            code = r[0].strip()
            name = r[1].strip()
            fields = [
                ("Leveranciersnummer", code),
                ("Leveranciersnaam", name),
                ("Straat", r[2].strip()),
                ("Huisnummer", r[3].strip()),
                ("Toevoeging huisnummer", r[4].strip()),
                ("Postcode", r[5].strip()),
                ("Plaats", r[6].strip()),
                ("Landcode", r[7].strip()),
                ("Telefoon", r[8].strip()),
                ("Fax", r[9].strip()),
                ("Email", r[10].strip()),
                ("BTW Nummer", r[11].strip()),
                ("Leveranciersgroep", r[12].strip()),
                ("Betalingsconditie", r[13].strip()),
                ("Financiële groep", r[14].strip()),
                ("Franco orderbedrag", r[15].strip()),
                ("Opgave levertijd", r[16].strip()),
                ("Valutacode", r[17].strip()),
            ]
            rows.append((name, code, fields))
    return rows

def create_manufacturers_workbook():
    import os
    import requests
    import json
    # Step 1: parse local manufacturers from lev file
    rows = process_lev_file()
    local_names = set(name for name, code, fields in rows)

    # Step 2: fetch all manufacturers from KatanaPIM
    url = "https://img.katanapim.com/api/v2/manufacturers"
    headers = {
        "apikey": "69d88bd8-08b6-4885-bcdc-e900c7b564c5",
        "Content-Type": "application/json"
    }
    try:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        all_manufs = resp.json()
    except Exception as e:
        print(f"Failed to fetch manufacturers from KatanaPIM: {e}")
        return
    # Build name→manufacturer object map
    name_to_manuf = {}
    if isinstance(all_manufs, list):
        for m in all_manufs:
            if m.get("name"):
                name_to_manuf[m["name"]] = m
    elif isinstance(all_manufs, dict) and "items" in all_manufs:
        for m in all_manufs["items"]:
            if m.get("name"):
                name_to_manuf[m["name"]] = m
    else:
        print("Could not parse manufacturers API response")
        return

    # Step 3: determine which names exist remotely but not locally
    remote_names = set(name_to_manuf.keys())
    missing_names = sorted(remote_names - local_names)

    # Step 4: prepare log file
    log_dir = os.path.join(BASE_DIR, "converted-files")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "log-manufacturer.txt")
    with open(log_path, "w", encoding="utf-8") as logf:
        logf.write("")  # clear

    # Step 5: for each missing, unpublish via PUT
    unpublish_count = 0
    for name in missing_names:
        m = name_to_manuf[name]
        manuf_id = m.get("id")
        if manuf_id is None:
            continue
        # Build payload: preserve code/name/description/displayOrder, set published=False
        payload = {
            "code": m.get("code"),
            "name": m.get("name"),
            "description": m.get("description"),
            "displayOrder": m.get("displayOrder"),
            "published": False
        }
        put_url = f"https://img.katanapim.com/api/v2/manufacturers/{manuf_id}"
        try:
            put_resp = requests.put(put_url, headers=headers, json=payload)
            ok = put_resp.status_code in (200, 204)
        except Exception as e:
            ok = False
            put_resp = None
        with open(log_path, "a", encoding="utf-8") as logf:
            if ok:
                logf.write(f"Unpublished manufacturer: {name} (id={manuf_id}, code={m.get('code')})\n")
                unpublish_count += 1
            else:
                logf.write(f"FAILED to unpublish {name} (id={manuf_id}, code={m.get('code')}), status={getattr(put_resp,'status_code',None)}\n")

    print(f"Unpublished {unpublish_count} manufacturers (not in lev file). Log: {log_path}")
    return

def run_conversion(suffix, folder, formula, debug_artikelnummer=None):
    import glob
    global INPUT_PATH, ATK_INPUT_PATH, MVM_INPUT_PATH, UIC_INPUT_PATH, STC_INPUT_PATH
    global OUTPUT_PATH, GP_OUTPUT_PATH, GLOBAL_FORMULA, GLOBAL_SUFFIX
    global LOG_FILE, ATK_LOG_FILE
    INPUT_PATH = find_file_by_prefix(suffix, "art")
    ATK_INPUT_PATH = find_file_by_prefix(suffix, "atk")
    MVM_INPUT_PATH = find_file_by_prefix(suffix, "mvm")
    UIC_INPUT_PATH = find_file_by_prefix(suffix, "uic") or ""
    STC_INPUT_PATH = find_file_by_prefix(suffix, "stc") or ""
    OUTPUT_PATH = os.path.join(BASE_DIR, "converted-files", folder, f"art0000{suffix}.xlsx")
    GP_OUTPUT_PATH = os.path.join(BASE_DIR, "converted-files", folder, f"art0000{suffix}_GP.xlsx")
    GLOBAL_FORMULA = formula
    GLOBAL_SUFFIX = suffix
    # per-suffix logbestanden
    LOG_FILE     = os.path.join(BASE_DIR, "converted-files", folder, f"log_{suffix}.txt")
    ATK_LOG_FILE = os.path.join(BASE_DIR, "converted-files", folder, f"ATK-log_{suffix}.txt")
    # ensure log directory exists before clearing logs
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    os.makedirs(os.path.dirname(ATK_LOG_FILE), exist_ok=True)
    # clear eerdere logs
    open(LOG_FILE,     "w", encoding="utf-8").close()
    open(ATK_LOG_FILE, "w", encoding="utf-8").close()

    # ensure catalogus uitvoering & soort are synced for this suffix
    process_uic_and_stc()

    art = process_art_data()
    if not art:
        return
    # Filter art list if debug_artikelnummer is set
    if debug_artikelnummer:
        art = [r for r in art if r[0].strip() == debug_artikelnummer.strip()]
    atk, uniq = process_atk_data()
    extra_atk = determine_atk_fields(uniq, atk)
    create_art_workbook(art, atk, extra_atk)
    mvm = process_mvm_file()
    create_gp_workbook(art, mvm)

def export_image_mapping_to_xlsx():
    """
    Zoek naar een bestand dat begint met adc en eindigt op in (adc*in),
    lees het bestand, en exporteer per unieke SKU een mapping van DisplayOrder naar image path
    naar images_in.xlsx (SKU, Name, Image_00, Image_01, ...)
    """
    import glob
    import openpyxl
    import csv
    import os
    # Zoek adc*in bestand in LOCAL_INPUT_DIR
    adc_pattern = os.path.join(LOCAL_INPUT_DIR, "adc*in")
    adc_files = glob.glob(adc_pattern)
    if not adc_files:
        print("Geen adc*in bestand gevonden voor image-mapping.")
        return
    adc_file = adc_files[0]
    print(f"Image-mapping: gebruik adc-bestand: {adc_file}")

    # Zoek bijbehorend art*in bestand (zelfde suffix als adc_file)
    art_file = None
    adc_basename = os.path.basename(adc_file)
    suffix_match = None
    import re
    m = re.match(r"adc(\d+)in", adc_basename)
    if m:
        suffix_match = m.group(1)
    # Fallback: pak eerste art*in bestand
    if suffix_match:
        art_pattern = os.path.join(LOCAL_INPUT_DIR, f"art{suffix_match}in")
        art_candidates = glob.glob(art_pattern)
        if art_candidates:
            art_file = art_candidates[0]
    if not art_file:
        # fallback: zoek eerste art*in bestand
        art_pattern = os.path.join(LOCAL_INPUT_DIR, "art*in")
        art_candidates = glob.glob(art_pattern)
        if art_candidates:
            art_file = art_candidates[0]
    sku_name_map = {}
    if art_file:
        with open(art_file, "r", encoding="latin1", errors="ignore") as af:
            reader = csv.reader(af, delimiter=",", quotechar='"')
            for row in reader:
                if len(row) >= 2:
                    sku = row[0].strip()
                    name = row[1].strip()
                    if sku:
                        sku_name_map[sku] = name
    # Lees adc data en bouw mapping
    sku_map = {}
    max_images = 0
    with open(adc_file, "r", encoding="latin1", errors="ignore") as f:
        reader = csv.reader(f, delimiter=",", quotechar='"')
        for row in reader:
            if len(row) < 5:
                continue
            sku = row[0].strip()
            displayorder = row[3].strip()
            img_path = row[4].strip() if len(row) > 4 and row[4].strip() else (row[2].strip() if len(row) > 2 else "")
            if not sku or not displayorder or not img_path:
                continue
            # Always extract filename from img_path
            filename = os.path.basename(img_path)
            url = f"https://img.leenweb.app/katanaPIM/dropfiles/output_files/ART_DOCS/{filename}"
            sku_map.setdefault(sku, {})
            sku_map[sku][displayorder.zfill(2)] = url
    # Bepaal het maximale aantal images per SKU
    for mapping in sku_map.values():
        n = len(mapping)
        if n > max_images:
            max_images = n
    # Verzamel alle unieke displayorders (gesorteerd, als '00', '01', ...)
    all_displayorders = set()
    for mapping in sku_map.values():
        all_displayorders.update(mapping.keys())
    displayorders_sorted = sorted(all_displayorders)
    # Bouw headers: SKU, Name, Image_00, Image_01, ...
    headers = ["SKU", "Name"] + [f"Image_{d}" for d in displayorders_sorted]
    # Maak workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for sku, mapping in sku_map.items():
        name = sku_name_map.get(sku, sku)
        row = [sku, name]
        for d in displayorders_sorted:
            row.append(mapping.get(d, ""))
        ws.append(row)
    # Schrijf weg naar images_in.xlsx in BASE_DIR/converted-files/in/
    out_dir = os.path.join(BASE_DIR, "converted-files", "in")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "images_in.xlsx")
    wb.save(out_path)
    print(f"Image mapping geëxporteerd naar: {out_path}")


def main():
    # Debug: only process a specific artikelnummer if set
    DEBUG_ARTIKELNUMMER = ""  # <-- pas deze waarde aan voor andere artikelen
    # Map folder and formula per suffix type
    folder_formula_map = {
        "pw": "Pronto Wonen",
        "bs": "Baenks",
        "in": "INHOUSE",
        "pm": "Profijt Meubel"
    }
    processed_suffixes = []
    for suffix in detect_suffixes():
        folder = suffix[-2:]  # 'in', 'pm', etc.
        formula = folder_formula_map.get(folder, "Onbekend")
        run_conversion(suffix, folder, formula, DEBUG_ARTIKELNUMMER)
        processed_suffixes.append((suffix, folder, formula))
        if folder == "in":
            export_image_mapping_to_xlsx()

    create_manufacturers_workbook()

    # --- Verstuur e-mailmelding (HTML-transcriptie) ---
    EMAIL_HTML_PATH = os.path.join(BASE_DIR, "converted-files", "email-na-conversie.html")
    # Dynamisch email_html opbouwen per verwerkte suffix (verbeterde HTML-versie)
    email_html = """<!DOCTYPE html>
<html lang="nl">
<head>
  <meta charset="UTF-8">
  <title>KatanaPIM Export</title>
</head>
<body>
<div style="max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 5px 15px rgba(0,0,0,0.05);">
  <div style="background-color: #007aeb; color: white; text-align: center; padding: 30px 20px;">
    <h1 style="margin: 0; font-size: 20px;">KatanaPIM Export Gereed</h1>
  </div>
  <div style="padding: 30px 20px;">
    <h2 style="font-size: 16px; margin-bottom: 20px;">Hey IMG,</h2>
    <p>De exports staan klaar om geïmporteerd te worden:</p>
"""

    for suffix, folder, formula in processed_suffixes:
        output_dir = os.path.join(BASE_DIR, "converted-files", folder)
        art_file = os.path.join(output_dir, f"art0000{suffix}.xlsx")
        gp_file = os.path.join(output_dir, f"art0000{suffix}_GP.xlsx")
        # Zoek JSON memovelden bestand (oude naam)
        memovelden_json = os.path.join(output_dir, f"art0000{suffix}_GP-memovelden.json")
        # Nieuwe naam zonder 0000
        memovelden_json_new = os.path.join(output_dir, f"art{suffix}_GP-memovelden.json")
        if os.path.exists(art_file) and os.path.exists(gp_file):
            url_prefix = f"https://img.leenweb.app/output_files/{folder}"
            email_html += f"""
    <hr>
    <h3 style="font-size: 14px; color: #007aeb;">— {formula.upper()} —</h3>
    <table style="width:100%; border-collapse:collapse; margin-bottom:20px;">
      <tr style="background-color:#f4f4f4; border:1px solid #ddd;">
        <td style="padding:10px; border:1px solid #ddd;"><strong>1 - Hoofdstructuur</strong></td>
        <td style="padding:10px; border:1px solid #ddd;">
          <a href="{url_prefix}/art0000{suffix}.xlsx" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download bestand</a>
          <button onclick="navigator.clipboard.writeText('{url_prefix}/art0000{suffix}.xlsx')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
        </td>
      </tr>
      <tr style="background-color:#ffffff; border:1px solid #ddd;">
        <td style="padding:10px; border:1px solid #ddd;"><strong>2 - Hoofdartikelen</strong></td>
        <td style="padding:10px; border:1px solid #ddd;">
          <a href="{url_prefix}/art0000{suffix}_GP.xlsx" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download bestand</a>
          <button onclick="navigator.clipboard.writeText('{url_prefix}/art0000{suffix}_GP.xlsx')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
        </td>
      </tr>
      <tr style="background-color:#f4f4f4; border:1px solid #ddd;">
        <td style="padding:10px; border:1px solid #ddd;"><strong>3 - Artikelen</strong></td>
        <td style="padding:10px; border:1px solid #ddd;">
          <a href="{url_prefix}/art0000{suffix}.xlsx" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download bestand</a>
          <button onclick="navigator.clipboard.writeText('{url_prefix}/art0000{suffix}.xlsx')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
        </td>
      </tr>
      <tr style="background-color:#ffffff; border:1px solid #ddd;">
        <td style="padding:10px; border:1px solid #ddd;"><strong>4 - Memovelden GrandParents</strong></td>
        <td style="padding:10px; border:1px solid #ddd;">
"""
            # Prefer new JSON name, fallback to old if not present
            if os.path.exists(memovelden_json_new):
                email_html += f"""
          <a href="{url_prefix}/art{suffix}_GP-memovelden.json" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download bestand</a>
          <button onclick="navigator.clipboard.writeText('{url_prefix}/art{suffix}_GP-memovelden.json')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
"""
            elif os.path.exists(memovelden_json):
                email_html += f"""
          <a href="{url_prefix}/art{suffix}_GP-memovelden.json" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download bestand</a>
          <button onclick="navigator.clipboard.writeText('{url_prefix}/art{suffix}_GP-memovelden.json')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
"""
            email_html += """
        </td>
      </tr>
    </table>
"""

    email_html += """
    <hr>
    <h3 style="font-size: 14px; color: #007aeb;">— Log bestanden —</h3>
"""
    log_url = "https://img.leenweb.app/output_files/logs.zip"
    email_html += f'''
      <p>
        <a href="{log_url}" style="display:inline-block;padding:10px 16px;background-color:#007aeb;color:#fff;text-decoration:none;border-radius:5px;">Download log bestanden</a>
        <button onclick="navigator.clipboard.writeText('{log_url}')" style="margin-left:10px; padding:10px 16px; background-color:#eeeeee; color:#333; border:none; border-radius:5px; cursor:pointer;">📋 Kopieer link</button>
      </p>
    '''
    email_html += """    <p style="margin-top: 20px;">Succes met de import!</p>
  </div>
  <div style="font-size: 12px; color: #888; text-align: center; padding: 20px;">&copy; 2025 Leenweb B.V.. Alle rechten voorbehouden.</div>
</div>
<script>
// Clipboard support for fallback (optioneel, moderne clients ondersteunen navigator.clipboard)
// Dit script is optioneel en voegt geen extra functionaliteit toe als navigator.clipboard werkt.
// Kan worden uitgebreid voor fallback.
</script>
</body>
</html>"""

    try:
        with open(EMAIL_HTML_PATH, "w", encoding="utf-8") as f:
            f.write(email_html)
        print(f"✉️  E-mailbestand aangemaakt op: {EMAIL_HTML_PATH}")
    except Exception as e:
        print(f"Fout bij aanmaken e-mailbestand: {e}")

    # --- Maak logs.zip met alle logbestanden ---
    import zipfile
    log_zip_path = os.path.join(BASE_DIR, "converted-files", "logs.zip")
    with zipfile.ZipFile(log_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for suffix, folder, _ in processed_suffixes:
            folder_path = os.path.join(BASE_DIR, "converted-files", folder)
            for fname in [f"log_{suffix}.txt", f"ATK-log_{suffix}.txt"]:
                file_path = os.path.join(folder_path, fname)
                if os.path.exists(file_path):
                    arcname = os.path.join(folder, fname)
                    zipf.write(file_path, arcname=arcname)
        # voeg ook log-manufacturer.txt toe
        manufacturer_log = os.path.join(BASE_DIR, "converted-files", "log-manufacturer.txt")
        if os.path.exists(manufacturer_log):
            zipf.write(manufacturer_log, arcname="log-manufacturer.txt")

    # --- Verstuur e-mail met HTML-inhoud ---
    try:
        with open(EMAIL_HTML_PATH, "r", encoding="utf-8") as f:
            email_content = f.read()

        msg = EmailMessage()
        msg['Subject'] = 'KatanaPIM Export Gereed'
        msg['From'] = 'noreply@leenweb.nl'
        msg['To'] = 'martijn@leenweb.nl'
        msg.set_content("De HTML-weergave van deze e-mail is vereist.")
        msg.add_alternative(email_content, subtype='html')

        with smtplib.SMTP('localhost') as smtp:
            smtp.send_message(msg)
        print("📧 E-mail verzonden naar martijn@leenweb.nl")
    except Exception as e:
        print(f"Fout bij verzenden e-mail: {e}")

if __name__=="__main__":
    main()